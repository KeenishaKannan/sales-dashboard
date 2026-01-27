# app.py
import streamlit as st
import pandas as pd
import plotly.express as px
from pathlib import Path
import re

st.set_page_config(page_title="CAL Sales Trend Analysis Dashboard", layout="wide")
st.title("CAL Sales Trend Analysis Dashboard")

EXCEL_FILE = "CAL Sales Data for Dashboard.xlsx"

# =========================
# Helpers
# =========================
def is_month(x):
    try:
        v = int(float(x))
        return 1 <= v <= 12
    except:
        return False

def to_int(x):
    try:
        return int(float(x))
    except:
        return None

def clean_series_name(s):
    if pd.isna(s):
        return None
    s = str(s)
    s = re.sub(r"[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]", "", s)
    s = re.sub(r"\|\s*nan", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\bnan\b", "", s, flags=re.IGNORECASE)
    return s.replace("|", "").strip()

def find_header_rows(df):
    for i in range(min(40, len(df))):
        if sum(is_month(v) for v in df.iloc[i]) >= 6:
            return i, i - 1 if i > 0 else None
    return None, None

def build_long_format(raw_df, sheet_name):
    header_row, year_row = find_header_rows(raw_df)
    if header_row is None:
        return None, f"Could not detect month headers in sheet '{sheet_name}'."

    months = raw_df.iloc[header_row]
    years = raw_df.iloc[year_row] if year_row is not None else [""] * len(months)

    year_ff, current_year = [], None
    for y in years:
        yi = to_int(y)
        if yi:
            current_year = yi
        year_ff.append(current_year)

    col_dates = {}
    for idx in range(len(months)):
        m, y = to_int(months[idx]), year_ff[idx]
        if m and y:
            col_dates[idx] = pd.Timestamp(year=y, month=m, day=1)

    id_cols = [i for i in range(len(months)) if i not in col_dates]
    data = raw_df.iloc[header_row + 1:].reset_index(drop=True)

    records = []
    for r in range(len(data)):
        label_parts = []
        for i in id_cols:
            v = data.iat[r, i]
            if pd.notna(v) and str(v).strip():
                label_parts.append(str(v).strip())

        series = clean_series_name(" ".join(label_parts))
        if not series:
            continue

        # 🔥 FIX — remove TOTAL rows (any language, any hidden chars)
        series_clean = series.lower().replace(" ", "")
        if ("total" in series_clean) or ("総計" in series_clean) or ("合計" in series_clean):
            continue

        for c, d in col_dates.items():
            val = pd.to_numeric(data.iat[r, c], errors="coerce")
            if pd.notna(val):
                records.append((series, d, round(float(val))))

    return pd.DataFrame(records, columns=["Series", "Date", "Value"]), None


def standardize_tidy_amount(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if "Amount" not in d.columns and "Amount (MY)" in d.columns:
        d["Amount"] = pd.to_numeric(d["Amount (MY)"], errors="coerce")
    else:
        if "Amount" in d.columns:
            d["Amount"] = pd.to_numeric(d["Amount"], errors="coerce")
    return d

# =========================
# Load Excel
# =========================
if not Path(EXCEL_FILE).exists():
    st.error("Excel file not found")
    st.stop()

xls = pd.ExcelFile(EXCEL_FILE)
sheet_names = xls.sheet_names

items_sheet = "ITEMS"
customers_sheet = "CUSTOMERS"

raw_items = pd.read_excel(EXCEL_FILE, sheet_name=items_sheet, header=None)
raw_customers = pd.read_excel(EXCEL_FILE, sheet_name=customers_sheet, header=None)

items_df, _ = build_long_format(raw_items, items_sheet)
customers_df, _ = build_long_format(raw_customers, customers_sheet)

# =========================
# Overall Trend
# =========================
st.markdown("---")
st.header("Overall Monthly Sales Trends")

c_items, c_customers = st.columns(2)

# -------- Items --------
with c_items:
    st.subheader("Trend Analysis of Items Based on Sales Amount")
    st.caption("💡 Shows the total monthly sales amount (MYR) summed across all items to track overall sales performance over time.")


    items_total = items_df.groupby("Date")["Value"].sum().reset_index()

    fig_items = px.line(
        items_total,
        x="Date",
        y="Value",
        markers=True,
        labels={"Value": "Sales Amount (MYR)"}
    )

    fig_items.update_traces(
        mode="lines+markers",
        hovertemplate="%{x|%b %Y}<br>Sales=%{y:,.0f}<extra></extra>"
    )

    fig_items.update_layout(
        template="plotly_white",
        height=420
    )

    st.plotly_chart(fig_items, use_container_width=True)

# -------- Customers --------
with c_customers:
    st.subheader("Trend Analysis of Customers Based on Items Sold")
    st.caption("💡 Shows the total monthly quantity sold summed across all customers to track overall demand volume over time")


    cust_total = customers_df.groupby("Date")["Value"].sum().reset_index()

    fig_cust = px.line(
        cust_total,
        x="Date",
        y="Value",
        markers=True,
        labels={"Value": "Quantity Sold"}
    )

    fig_cust.update_traces(
        mode="lines+markers",
        hovertemplate="%{x|%b %Y}<br>Quantity=%{y:,.0f}<extra></extra>"
    )

    fig_cust.update_layout(
        template="plotly_white",
        height=420
    )

    st.plotly_chart(fig_cust, use_container_width=True)

# =========================
# Year-to-Year Sales Comparison
# =========================
st.markdown("---")
st.header("Year-to-Year Sales Comparison")

metric = st.radio(
    "Select metric",
    ["Sales Amount (MYR)", "Quantity Sold"],
    horizontal=True,
)

base_df = items_df if metric == "Sales Amount (MYR)" else customers_df

entity_label = "Item" if metric == "Sales Amount (MYR)" else "Customer"
entities = sorted(base_df["Series"].unique())

selected_entities = st.multiselect(
    f"Filter by {entity_label}",
    entities,
    default=[],
    key="yoy_optional_entity_filter"
)

df_yoy = base_df.copy()
df_yoy["Date"] = pd.to_datetime(df_yoy["Date"])
df_yoy["Year"] = df_yoy["Date"].dt.year
df_yoy["MonthNum"] = df_yoy["Date"].dt.month
df_yoy["Month"] = df_yoy["Date"].dt.strftime("%b")

if selected_entities:
    df_yoy = df_yoy[df_yoy["Series"].isin(selected_entities)]

years = sorted(df_yoy["Year"].unique())
year_range = st.slider(
    "Select year range",
    min_value=int(min(years)),
    max_value=int(max(years)),
    value=(int(min(years)), int(max(years))),
    step=1
)

df_yoy = df_yoy[
    (df_yoy["Year"] >= year_range[0]) &
    (df_yoy["Year"] <= year_range[1])
]

yoy_df = (
    df_yoy
    .groupby(["Year", "MonthNum", "Month"], as_index=False)["Value"]
    .sum()
)

month_order = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
yoy_df["Month"] = pd.Categorical(yoy_df["Month"], categories=month_order, ordered=True)
yoy_df = yoy_df.sort_values(["Year", "MonthNum"])

fig = px.line(
    yoy_df,
    x="Month",
    y="Value",
    color="Year",
    markers=True,
    labels={
        "Value": metric,
        "Month": "Month",
        "Year": "Year",
    },
)

fig.update_traces(
    hovertemplate=
        "Year=%{fullData.name}<br>"
        "Month=%{x}<br>"
        f"{metric}=%{{y:,.0f}}<extra></extra>"
)

fig.update_layout(
    template="plotly_white",
    xaxis=dict(categoryorder="array", categoryarray=month_order),
    yaxis_title=metric,
    hovermode="x unified",
    height=520
)

st.plotly_chart(fig, use_container_width=True)


# =========================
# Pie Charts
# =========================
st.markdown("---")
st.header("Monthly Sales Breakdown")
st.caption("💡 Select a month to view the breakdown for Items (MYR) and Customers (Quantity Sold).")

items_df["Date"] = pd.to_datetime(items_df["Date"])
customers_df["Date"] = pd.to_datetime(customers_df["Date"])

all_months = sorted(
    set(items_df["Date"].dt.to_period("M")).union(set(customers_df["Date"].dt.to_period("M")))
)
month_labels = [m.to_timestamp().strftime("%b %Y") for m in all_months]

selected_month_label = st.selectbox(
    "Select month",
    month_labels,
    index=len(month_labels) - 1 if len(month_labels) else 0,
    key="pie_month_select"
)
selected_month = pd.to_datetime(selected_month_label, format="%b %Y").to_period("M")

TOP_N_ITEMS = 10
TOP_N_CUSTOMERS = 10

def top_n_with_others(df, name_col, value_col, top_n, others_label="Others"):
    d = (
        df.groupby(name_col, as_index=False)[value_col]
        .sum()
        .sort_values(value_col, ascending=False)
        .reset_index(drop=True)
    )
    if len(d) <= top_n:
        return d

    top = d.iloc[:top_n].copy()
    others_val = d.iloc[top_n:][value_col].sum()
    others_row = pd.DataFrame([{name_col: others_label, value_col: others_val}])
    out = pd.concat([top, others_row], ignore_index=True)

    out = out.sort_values(value_col, ascending=False).reset_index(drop=True)
    return out

c1, c2 = st.columns(2)

# -------------------------
# Pie 1: Sales Based on Items (MYR)
# -------------------------
with c1:
    st.subheader("Sales Based on Items (MYR)")

    # Filter for selected month
    items_m = items_df[items_df["Date"].dt.to_period("M") == selected_month].copy()
    items_m["Series"] = items_m["Series"].astype(str).str.strip()

    # REMOVE invalid entries (0, numeric-only)
    items_m = items_m[
        (items_m["Series"] != "") &
        (~items_m["Series"].str.fullmatch(r"0|0\.0|0\.00")) &
        (~items_m["Series"].str.fullmatch(r"\d+(\.\d+)?"))
    ].reset_index(drop=True)

    # Aggregate totals per item
    items_pie = (
        items_m.groupby("Series", as_index=False)["Value"]
        .sum()
        .sort_values("Value", ascending=False)   
        .reset_index(drop=True)
    )

    # Top-N + Others
    if len(items_pie) > TOP_N_ITEMS:
        top = items_pie.iloc[:TOP_N_ITEMS].copy()
        others_val = items_pie.iloc[TOP_N_ITEMS:]["Value"].sum()
        others_row = pd.DataFrame([{"Series": "Others", "Value": others_val}])
        items_pie = pd.concat([top, others_row], ignore_index=True)

    # -------------------------
    # Fixed color map
    # -------------------------
    item_colors = {
        "Strawberry": "pink",       
        "Tomato": "red",            
        "Sweet Corn": "yellow",
        "Daikon (Raddish)": "#26A69A",
        "Spinach": "#4CAF50",
        "Tong Hou": "#FFA726",
        "Seedlings": "#9575CD",
        "Cabbage": "#CFD8DC",
        "Asparagus": "#1E88E5",
        "Shiro Negi": "#546E7A",
        "Others": "gray"
    }

    # -------------------------
    # Pie chart
    # -------------------------
    fig_items_pie = px.pie(
        items_pie,
        names="Series",
        values="Value",
        color="Series",
        color_discrete_map=item_colors
    )

    fig_items_pie.update_traces(
        direction="clockwise",          
        sort=False,                     
        rotation=0,
        textinfo="percent",
        textposition="inside",
        hovertemplate="Item=%{label}<br>Sales=%{value:,.0f} MYR<br>%{percent}<extra></extra>"
    )

    fig_items_pie.update_layout(
        template="plotly_white",
        height=520,
        legend_title_text="Item"
    )

    st.plotly_chart(fig_items_pie, use_container_width=True)



with c2:
    st.subheader("Sales Based on Supermarkets (RTL - Sales Amount)")

    tidy_df = pd.read_excel(EXCEL_FILE, sheet_name="Tidy Data")
    tidy_df = standardize_tidy_amount(tidy_df)

    tidy_df["Date"] = pd.to_datetime(
        tidy_df["Year"].astype(str) + "-" +
        tidy_df["Month"].astype(str).str.zfill(2) + "-01"
    )

    rtl_df = tidy_df[tidy_df["Section"].astype(str).str.upper() == "RTL"].copy()

    rtl_month = rtl_df[rtl_df["Date"].dt.to_period("M") == selected_month]

    if rtl_month.empty:
        latest_month = rtl_df["Date"].max().to_period("M")
        rtl_month = rtl_df[rtl_df["Date"].dt.to_period("M") == latest_month]

    rtl_pie = (
        rtl_month
        .groupby("Sub section", as_index=False)["Amount"]
        .sum()
        .rename(columns={"Sub section": "Supermarket", "Amount": "Sales"})
        .sort_values("Sales", ascending=False)
    )

    rtl_colors = {
        "AEN": "#636EFA",
        "JGC": "#EF553B",
        "TFP": "#00CC96",
        "IST": "#AB63FA",
        "QRA": "#19D3F3",
        "MCV": "#FFA15A",
        "OTH": "#9E9E9E",
    }

    fig_rtl = px.pie(
        rtl_pie,
        names="Supermarket",
        values="Sales",
        color="Supermarket",
        color_discrete_map=rtl_colors,
    )

    fig_rtl.update_traces(
        direction="clockwise",
        sort=False,
        rotation=0,
        textinfo="percent",
        textposition="inside",
    )

    fig_rtl.update_layout(
        template="plotly_white",
        height=520,
        legend_title_text="Supermarket",
    )

    st.plotly_chart(fig_rtl, use_container_width=True)

# =========================
# Heatmap: Item x Customer (Monthly Sales)
# =========================
st.subheader("Item × Customer Sales Heatmap")
st.caption("💡 This heatmap visualizes sales intensity (MYR) across items and customers to identify key sales drivers.")

tidy_df = pd.read_excel(EXCEL_FILE, sheet_name="Tidy Data")
tidy_df = standardize_tidy_amount(tidy_df)

tidy_df["Date"] = pd.to_datetime(
    tidy_df["Year"].astype(str) + "-" +
    tidy_df["Month"].astype(int).astype(str).str.zfill(2) + "-01",
    errors="coerce"
)

months = sorted(tidy_df["Date"].dropna().dt.to_period("M").unique())
month_labels = [m.to_timestamp().strftime("%b %Y") for m in months]

selected_month_label = st.selectbox(
    "Select month",
    month_labels,
    index=len(month_labels) - 1 if month_labels else 0,
    key="heatmap_month_select"
)

selected_month = pd.to_datetime(
    selected_month_label, format="%b %Y"
).to_period("M")

d = tidy_df[tidy_df["Date"].dt.to_period("M") == selected_month].copy()

d["Item"] = d["Item"].astype(str)
d["Sales to"] = d["Sales to"].astype(str)

heat_df = (
    d.groupby(["Item", "Sales to"], as_index=False)["Amount"]
    .sum()
)

heat_pivot = (
    heat_df
    .pivot(index="Item", columns="Sales to", values="Amount")
    .fillna(0)
)

if heat_pivot.empty:
    st.warning("No data available for the selected month.")
else:
    cols_to_drop = [
        c for c in heat_pivot.columns
        if str(c).strip().upper() == "CAI"
        or "OTHER" in str(c).upper()
        or "E-COMMERCE" in str(c).upper()
        or "ECOMMERCE" in str(c).upper()
        or "E COMM" in str(c).upper()
        or "E-COMM" in str(c).upper()
    ]

    if cols_to_drop:
        heat_pivot = heat_pivot.drop(columns=cols_to_drop, errors="ignore")

    heat_pivot = heat_pivot.loc[
        heat_pivot.sum(axis=1).sort_values(ascending=False).index,
        heat_pivot.sum(axis=0).sort_values(ascending=False).index
    ]

    vmax_default = int(heat_pivot.values.max())

    vmax = st.slider(
        "Adjust colour range (MYR)",
        min_value=0,
        max_value=vmax_default,
        value=vmax_default,
        key="heatmap_colour_range"
    )

    fig = px.imshow(
        heat_pivot,
        aspect="auto",
        color_continuous_scale="Blues",
        zmin=0,
        zmax=vmax,
        labels={
            "x": "Customer",
            "y": "Item",
            "color": "Sales (MYR)"
        }
    )

    fig.update_layout(
        template="plotly_white",
        height=620,
        xaxis_tickangle=-45
    )

    fig.update_traces(
        hovertemplate=
        "Item=%{y}<br>"
        "Customer=%{x}<br>"
        "Sales=%{z:,.0f} MYR<extra></extra>"
    )

    st.plotly_chart(fig, use_container_width=True)

# ======================================================
# Sales per Retails (RTL)
# ======================================================
st.header("Sales per Retail")
st.caption("💡 Monthly sales trend for major retailers.")

import openpyxl
import pandas as pd
import plotly.express as px
import re

def load_rtl_long(excel_file: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb["RTL"]

    YEAR_ROW = 2
    MONTH_ROW = 3
    FIRST_DATA_ROW = 4
    CUSTOMER_COL = 1

    year_cols = []
    max_col = ws.max_column

   
    for c in range(1, max_col + 1):
        v = ws.cell(YEAR_ROW, c).value
        if isinstance(v, (int, float)) and int(v) == v and 2000 <= int(v) <= 2100:
            year_cols.append((c, int(v)))

    if not year_cols:
        return pd.DataFrame(columns=["Date", "Retail", "Value"])

    
    col_to_date = {}
    for i, (start_c, year_val) in enumerate(year_cols):
        end_c = (year_cols[i + 1][0] - 1) if i + 1 < len(year_cols) else max_col

        for c in range(start_c, end_c + 1):
            m = ws.cell(MONTH_ROW, c).value
            if isinstance(m, (int, float)) and int(m) == m and 1 <= int(m) <= 12:
                col_to_date[c] = pd.Timestamp(year=year_val, month=int(m), day=1)

    if not col_to_date:
        return pd.DataFrame(columns=["Date", "Retail", "Value"])

   
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        retail = ws.cell(r, CUSTOMER_COL).value
        if retail is None or str(retail).strip() == "":
            continue

        retail = str(retail).strip()

        
        if re.fullmatch(r"\d+(\.\d+)?", retail):
            continue

        for c, dt in col_to_date.items():
            val = ws.cell(r, c).value
            if val is None or val == "":
                continue

            num = pd.to_numeric(val, errors="coerce")
            if pd.isna(num):
                continue

            rows.append((dt, retail, float(num)))

    out = pd.DataFrame(rows, columns=["Date", "Retail", "Value"])
    if out.empty:
        return out

    out = (
        out.groupby(["Date", "Retail"], as_index=False)["Value"]
           .sum()
           .sort_values(["Retail", "Date"])
    )
    return out



rtl_long = load_rtl_long(EXCEL_FILE)

if rtl_long.empty:
    st.error("No RTL data found after parsing (check RTL sheet layout).")
else:

    top_n = 8
    totals = rtl_long.groupby("Retail", as_index=False)["Value"].sum()
    top_retails = totals.sort_values("Value", ascending=False).head(top_n)["Retail"].tolist()
    rtl_plot = rtl_long[rtl_long["Retail"].isin(top_retails)].copy()

   
    fig_rtl = px.line(
        rtl_plot,
        x="Date",
        y="Value",
        color="Retail",
        markers=True, 
        labels={"Value": "Sales Amount (MYR)", "Date": "Month", "Retail": ""},
    )

    fig_rtl.update_traces(
        hovertemplate=
        "Month=%{x|%Y-%m}<br>"
        "Customer=%{fullData.name}<br>"
        "Sales Amount=%{y:,.0f}<extra></extra>"
    )

    fig_rtl.update_layout(
        legend_title_text="",
        yaxis_tickformat=",",
        xaxis_tickformat="%Y%m",
        height=520,
        margin=dict(l=40, r=40, t=20, b=40),
    )

    st.plotly_chart(fig_rtl, use_container_width=True)


# =========================
# Sales Change by Month – Items
# Sales by Customer Segment
# =========================
st.markdown("---")

col1, col2 = st.columns(2)

with col1:
    st.header("Sales Change by Month – Items")
    st.caption(
        "💡Shows how total monthly sales (MYR) change over time, broken down by key items, "
        "to highlight which items are driving overall sales growth or decline."
    )

  
    items_df = items_df[items_df["Series"].astype(str).str.strip().str.lower().ne("")]   
    items_df = items_df[items_df["Series"] != "0"]                                       
    items_df = items_df[~items_df["Series"].str.contains("^others ?0$", case=False)]     

    
    top_items = (
    items_df.groupby("Series")["Value"]
        .sum()
        .sort_values(ascending=False)
        .head(5)
        .index
     )

    items_df["Group"] = items_df["Series"].where(
        items_df["Series"].isin(top_items), "Others"
    )

    stack_items = (
        items_df.groupby(["Date", "Group"])["Value"]
        .sum()
        .reset_index()
    )

    group_order = [g for g in top_items if g != "Others"] + ["Others"]

    color_map = {
        "Strawberry": "pink",
        "Sweet Corn": "yellow",
        "Tomato": "red",
        "Daikon (Raddish)": "#26A69A",
        "Others": "gray",
    }

    fig_items = px.area(
        stack_items,
        x="Date",
        y="Value",
        color="Group",
        labels={"Value": "Sales Amount (MYR)"},
        category_orders={"Group": group_order},
        color_discrete_map=color_map,
    )

    fig_items.update_traces(
        hovertemplate="%{x|%b %Y}<br>%{y:,.0f}<extra></extra>"
    )

    fig_items.update_layout(
        height=520,
        legend_title_text="Group"
    )

    st.plotly_chart(fig_items, use_container_width=True)

# ======================================================
# Sales by Customer Segment
# ======================================================
with col2:
    st.header("Sales by Customer Segment")
    st.caption(
        "💡Shows the monthly sales breakdown (MYR) by customer segment, "
        "highlighting each segment’s contribution to total sales over time."
    )

    segment_sheet = "SEGMENT"

    raw_segment = pd.read_excel(
        EXCEL_FILE,
        sheet_name=segment_sheet,
        header=None
    )

    segment_df, segment_err = build_long_format(raw_segment, segment_sheet)

    if segment_df is None:
        st.error(segment_err)
    else:

        def extract_segment(name):
            name = name.upper()
            for seg in ["CAI", "DST", "HRC", "RTL", "OTH"]:
                if seg in name:
                    return seg
            return "OTH"

        segment_df["Segment"] = segment_df["Series"].apply(extract_segment)

        segment_monthly = (
            segment_df
            .groupby(["Date", "Segment"], as_index=False)["Value"]
            .sum()
            .round(0)
        )

        segment_colors = {
            "CAI": "#E53935",   
            "DST": "#90CAF9",   
            "HRC": "#1E3A8A",   
            "RTL": "#26A69A",   
            "OTH": "#9E9E9E"    
        }

        segment_order = ["CAI", "DST", "HRC", "RTL", "OTH"]

        fig_segment = px.bar(
            segment_monthly,
            x="Date",
            y="Value",
            color="Segment",
            barmode="stack",
            color_discrete_map=segment_colors,
            category_orders={"Segment": segment_order},
            labels={
                "Value": "Sales (MYR)",
                "Date": "Month"
            }
        )

        fig_segment.update_xaxes(
            tickformat="%Y-%m",
            tickangle=45
        )

        fig_segment.update_traces(
            hovertemplate=
            "Month=%{x|%Y-%m}<br>"
            "Segment=%{fullData.name}<br>"
            "Sales=%{y:,.0f}<extra></extra>"
        )

        fig_segment.update_layout(
            yaxis_tickformat=",",
            legend_title_text="Segment",
            height=520
        )

        st.plotly_chart(fig_segment, use_container_width=True)


# =========================
# Items & Customers Trend
# =========================
st.markdown("---")

c_items, c_customers = st.columns(2)

# -------- Items Trend --------
with c_items:
    simulate_st_header = st.header("Items Trend Analysis")
    st.caption(
        "💡 Shows the monthly sales trend (MYR) for selected items over time, "
        "allowing comparison of performance and seasonality across products."
    )

    items = sorted(items_df["Series"].unique())

    selected_items = st.multiselect(
        "Select items",
        items,
        default=items[:5],
        key="items_trend_select"
    )

    if selected_items:
        fig = px.line(
            items_df[items_df["Series"].isin(selected_items)],
            x="Date",
            y="Value",
            color="Series",
            markers=True,
            labels={"Value": "Sales Amount (MYR)"}
        )

        fig.update_traces(
            mode="lines+markers",
            hovertemplate="%{x|%b %Y}<br>%{y:,.0f}<extra></extra>"
        )

        st.plotly_chart(fig, use_container_width=True)


# -------- Customers Trend --------
with c_customers:
    st.header("Customers Trend Analysis")
    st.caption(
        "💡 Shows the monthly quantity sold to selected customers over time, "
        "allowing comparison of purchasing patterns and demand consistency."
    )

    
    customers_df["Series"] = (
        customers_df["Series"]
        .astype(str)
        .str.strip()
    )

   
    customers = sorted(
        customers_df.loc[
            (customers_df["Series"] != "") &
            (~customers_df["Series"].str.fullmatch(r"\d+(\.\d+)?"))
        ]["Series"].unique()
    )

    
    selected_customers = st.multiselect(
        "Select customers",
        customers,
        default=customers[:5],      
        key="customers_trend_select"
    )

   
    selected_customers = [
        c for c in selected_customers if c in customers
    ]

    if selected_customers:
        df_plot = customers_df[
            customers_df["Series"].isin(selected_customers)
        ]

        fig = px.line(
            df_plot,
            x="Date",
            y="Value",
            color="Series",
            markers=True,
            labels={"Value": "Quantity Sold"}
        )

        fig.update_traces(
            mode="lines+markers",
            hovertemplate="%{x|%b %Y}<br>%{y:,.0f}<extra></extra>"
        )

        st.plotly_chart(fig, use_container_width=True)


# =========================
# Growth & Action Insights
# =========================
st.markdown("---")
st.header("Growth & Action Insights")

import numpy as np

def _compute_mom_stats(df, entity_col, value_col="Value", months_window=6, cap_pct=300):
    d = df.copy()

    d["Date"] = pd.to_datetime(d["Date"])
    last_month = d["Date"].max()
    start_month = (last_month - pd.DateOffset(months=months_window - 1)).replace(day=1)
    d = d[d["Date"] >= start_month]

    m = (
        d.groupby([entity_col, "Date"])[value_col]
        .sum()
        .reset_index()
        .sort_values([entity_col, "Date"])
    )

    piv = m.pivot_table(index=entity_col, columns="Date", values=value_col, fill_value=0)

    med = piv.median(axis=1)
    thresh = (med * 0.05).clip(lower=1)

    mom_list = []
    cols = list(piv.columns)

    for i in range(1, len(cols)):
        prev = piv[cols[i - 1]]
        cur = piv[cols[i]]
        valid = prev >= thresh

        mom = pd.Series(np.nan, index=piv.index, dtype="float64")
        mom.loc[valid] = ((cur.loc[valid] - prev.loc[valid]) / prev.loc[valid]) * 100.0
        mom = mom.clip(lower=-cap_pct, upper=cap_pct)
        mom_list.append(mom.rename(cols[i]))

    mom_df = pd.DataFrame(mom_list).T

    avg_value_window = piv.mean(axis=1)
    avg_mom_window = mom_df.mean(axis=1, skipna=True)
    vol_mom_window = mom_df.std(axis=1, skipna=True)

    recent_cols = list(mom_df.columns)[-3:]
    avg_mom_recent = mom_df[recent_cols].mean(axis=1, skipna=True)

    out = pd.DataFrame({
        "Entity": mom_df.index,
        "avg_mom_recent": avg_mom_recent,
        "avg_mom_window": avg_mom_window,
        "vol_mom_window": vol_mom_window,
        "avg_value_window": avg_value_window,
    }).reset_index(drop=True)

    out["avg_mom_recent_display"] = out["avg_mom_recent"].fillna(0)
    out["avg_mom_window_display"] = out["avg_mom_window"].fillna(0)
    out["vol_mom_window_display"] = out["vol_mom_window"].fillna(0)

    return out

item_stats = _compute_mom_stats(items_df, entity_col="Series", value_col="Value", months_window=6, cap_pct=300)
cust_stats = _compute_mom_stats(customers_df, entity_col="Series", value_col="Value", months_window=6, cap_pct=300)

c1, c2 = st.columns(2)

with c1:
    st.subheader("Items With Rising Momentum")
    st.caption("💡 Months with Sales Increase: in the selected window (3/6/12 months), how many times did the item go up compared to the previous month.")

    months_window = st.radio(
        "Select time window",
        options=[3, 6, 12],
        index=2,
        horizontal=True,
        key="items_rising_window"
    )

    d = items_df.copy()
    d["Date"] = pd.to_datetime(d["Date"])

    last_month = d["Date"].max()
    start_month = (last_month - pd.DateOffset(months=months_window - 1)).replace(day=1)
    d = d[d["Date"] >= start_month]

    m = (
        d.groupby(["Series", "Date"], as_index=False)["Value"]
        .sum()
        .sort_values(["Series", "Date"])
    )

    def count_growth(g):
        return int((g["Value"].diff() > 0).sum())

    momentum = (
        m.groupby("Series", as_index=False)
        .apply(count_growth)
        .rename(columns={None: "Growth_Months"})
    )

    top_items = (
        momentum.sort_values(["Growth_Months", "Series"], ascending=[False, True])
        .head(10)
    )

    plot_df = top_items.sort_values("Growth_Months", ascending=True)

    fig = px.bar(
        plot_df,
        x="Growth_Months",
        y="Series",
        orientation="h",
        text="Growth_Months",
        labels={"Series": "Item", "Growth_Months": "Months with Sales Increase"},
    )

    fig.update_traces(
        marker_color="#22C55E",
        texttemplate="%{text}",
        textposition="outside",
        hovertemplate="Item=%{y}<br>Months with Increase=%{x}<extra></extra>",
        cliponaxis=False
    )

    fig.update_layout(
        template="plotly_white",
        height=520,
        showlegend=False,
        margin=dict(l=20, r=20, t=10, b=20),
        xaxis_title="Months with Sales Increase",
        yaxis_title="Item",
    )

    st.plotly_chart(fig, use_container_width=True)

with c2:
    st.subheader("Customer Sales Growth Trends")
    st.caption("💡 Customers categorized by improving or declining sales patterns over time.")

    months_window = st.radio(
        "Select time window",
        options=[3, 6, 12],
        index=1,
        horizontal=True,
        key="customer_growth_window"
    )

    d = customers_df.copy()
    d["Date"] = pd.to_datetime(d["Date"])

    last_month = d["Date"].max()
    start_month = (last_month - pd.DateOffset(months=months_window - 1)).replace(day=1)
    d = d[d["Date"] >= start_month]

    monthly = (
        d.groupby(["Series", "Date"])["Value"]
        .sum()
        .reset_index()
        .sort_values(["Series", "Date"])
    )

    def avg_monthly_change(x):
        return x["Value"].diff().mean()

    cs = (
        monthly.groupby("Series")
        .apply(avg_monthly_change)
        .reset_index(name="Avg_Change")
        .fillna(0)
    )

    def trend_label(v):
        if v > 0:
            return "Improving"
        if v < 0:
            return "Declining"
        return "Stable"

    cs["Trend"] = cs["Avg_Change"].apply(trend_label)

   
    cs = cs[(cs["Trend"] != "Stable") & (cs["Series"] != "0")]

    improving = cs[cs["Trend"] == "Improving"].sort_values("Avg_Change", ascending=False).head(6)
    declining = cs[cs["Trend"] == "Declining"].sort_values("Avg_Change").head(6)

    show = pd.concat([declining, improving]).sort_values("Avg_Change")

    fig_height = max(360, len(show) * 45)

    fig = px.bar(
        show,
        x="Avg_Change",
        y="Series",
        orientation="h",
        color="Trend",
        labels={
            "Series": "Customer",
            "Avg_Change": "Avg Monthly Sales Change (MYR)"
        },
        color_discrete_map={
            "Improving": "#2563EB",
            "Declining": "#EF4444"
        }
    )

    fig.update_traces(
        hovertemplate=
        "Customer=%{y}<br>"
        "Avg Monthly Change=%{x:,.0f} MYR"
        "<extra></extra>"
    )

    fig.update_layout(
        template="plotly_white",
        height=fig_height,
        xaxis_title="Avg Monthly Sales Change (MYR)",
        yaxis_title="Customer",
        legend_title_text="Trend",
        yaxis=dict(
            automargin=True,
            tickmode="linear"
        )
    )

    st.plotly_chart(fig, use_container_width=True)


# =========================
# Item Sales Distribution
# =========================
st.markdown("---")
st.header("Item Sales Distribution (Histogram + Bell Curve)")
st.caption(
    "💡 For the selected year, shows how many items fall into each average-monthly-sales range (MYR). "
    "Includes mean and median."
)

items_df["Year"] = items_df["Date"].dt.year
available_years = sorted(items_df["Year"].dropna().unique().tolist())

selected_year = st.selectbox(
    "Select year",
    available_years,
    index=len(available_years) - 1 if available_years else 0,
    key="hist_bell_year"
)

d = items_df[
    (items_df["Year"] == selected_year) &
    (items_df["Series"].str.lower() != "others")
].copy()

d["Date"] = pd.to_datetime(d["Date"])

monthly_item = (
    d.groupby(["Series", "Date"], as_index=False)["Value"]
    .sum()
)

item_avg = (
    monthly_item.groupby("Series", as_index=False)["Value"]
    .mean()
    .rename(columns={"Value": "AvgMonthlySales"})
)

vals = item_avg["AvgMonthlySales"].dropna().astype(float).values

if len(vals) < 3:
    st.warning("Not enough data to plot distribution for this year.")
else:
    import plotly.graph_objects as go

    bins = 15

    counts, edges = np.histogram(vals, bins=bins)
    bin_centers = (edges[:-1] + edges[1:]) / 2
    bin_width = edges[1] - edges[0]

    mu = float(np.mean(vals))
    sigma = float(np.std(vals, ddof=0)) if float(np.std(vals, ddof=0)) > 0 else 1.0
    median = float(np.median(vals))

    x = np.linspace(edges[0], edges[-1], 400)
    pdf = (1.0 / (sigma * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x - mu) / sigma) ** 2)
    bell_y = pdf * len(vals) * bin_width

    fig = go.Figure()

    fig.add_trace(
        go.Bar(
            x=bin_centers,
            y=counts,
            width=bin_width * 0.9,
            name="Number of items",
            hovertemplate=
            "Avg monthly sales ≈ %{x:,.0f} MYR<br>"
            "Items in this range = %{y}<extra></extra>"
        )
    )

    fig.add_trace(
        go.Scatter(
            x=x,
            y=bell_y,
            mode="lines",
            name="Overall distribution trend",
            hovertemplate="Sales = %{x:,.0f} MYR<extra></extra>"
        )
    )

    fig.add_vline(x=mu, line_dash="dash", annotation_text="Mean", annotation_position="top")
    fig.add_vline(x=median, line_dash="dot", annotation_text="Median", annotation_position="top")

    fig.update_layout(
        template="plotly_white",
        height=520,
        xaxis_title="Average Monthly Sales per Item (MYR)",
        yaxis_title="Number of Items",
        legend_title_text=""
    )

    st.plotly_chart(fig, use_container_width=True)

item_avg["AvgMonthlySales"] = item_avg["AvgMonthlySales"].round(0).astype(int)

c1, c2 = st.columns(2)

with c1:
    st.subheader("Top 10 items (highest avg monthly sales)")
    top10 = (
        item_avg.sort_values("AvgMonthlySales", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    top10.index = top10.index + 1
    st.dataframe(top10, use_container_width=True)

with c2:
    st.subheader("Bottom 10 items (lowest avg monthly sales)")
    bottom10 = (
        item_avg.sort_values("AvgMonthlySales", ascending=True)
        .head(10)
        .reset_index(drop=True)
    )
    bottom10.index = bottom10.index + 1
    st.dataframe(bottom10, use_container_width=True)

st.markdown("### 📊 Distribution Summary")

total_items = len(item_avg)
high_seller_threshold = 100_000
high_sellers = (item_avg["AvgMonthlySales"] >= high_seller_threshold).sum()

c1, c2, c3, c4 = st.columns(4)

c1.metric("Total Items", int(total_items))
c2.metric("Median Avg Monthly Sales (MYR) per item", f"{int(median):,}")
c3.metric("Mean Avg Monthly Sales (MYR) per item", f"{int(mu):,}")
c4.metric(
    "High-Selling Items",
    int(high_sellers),
    help=f"Items with ≥ {high_seller_threshold:,} MYR average monthly sales"
)
